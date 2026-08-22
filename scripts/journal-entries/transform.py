#!/usr/bin/env python3
"""
journal_entries 移行スクリプト（1/2: 変換）

「仕訳１を行ったリスト.xlsx」（取引先ごとの過去の仕訳結果、7シート）を読み込み、
Supabaseの journal_entries テーブルにそのまま投入できるJSON配列に変換してファイルへ書き出す。

このスクリプトはローカルのExcelファイルを直接読み込むだけで、Supabaseへの通信は一切行わない
（transform → import の2段階に分けているのは、変換結果を目で確認してからDBに書き込むため）。

出力JSONには個人の家計取引データ（店名・金額等）が含まれるため、
このリポジトリにはコミットしないこと（.gitignoreで除外済み）。

使い方:
  python3 transform.py <入力xlsxのパス> [出力jsonのパス]
  （出力を省略した場合は journal_entries.json という名前でカレントディレクトリに書き出す）

設計の詳細は gahoo-company リポジトリの
  product/journal-entries-db-design.md
を参照。日付の年推定・住友VISAのカード名義ヘッダー機構・列レイアウトの揺れ（みずほ銀行）などの
変換ルールはすべて、この design doc に記載の内容と実データ検証に基づく。
"""
import sys
import re
import json
import datetime

DASH_CHARS = set('ー－-−ｰ‐')


def is_dash_only(s):
    if not s:
        return False
    return all(ch in DASH_CHARS for ch in str(s))


def parse_month_day_text(s):
    m = re.match(r'^(\d{1,2})月(\d{1,2})日$', str(s).strip())
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def parse_full_date_text(s):
    """'2025/3/1' のような年月日まで含む表記（横浜VISAの一部行で使われる）"""
    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', str(s).strip())
    if not m:
        return None
    return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))


def to_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    s = s.replace('\\', '').replace(',', '').replace('¥', '')
    if s == '':
        return None
    return float(s) if '.' in s else int(s)


def resolve_date_same_year(month, day, ym_col):
    """銀行口座向け：月列(YYYYMM)の年をそのまま使う（月が一致することを実データで確認済み）"""
    year = int(ym_col) // 100
    return datetime.date(year, month, day)


def resolve_date_billing_lag(month, day, billing_ym):
    """カード明細向け：支払月列(YYYYMM)から、ラグ0〜3か月のいずれかで
    取引月と一致する方を年の決定に採用する（カードの締め日により1〜2か月ラグが基本、
    まれに0・3か月ラグのケースもあるため実データ検証で範囲を広げてある）"""
    billing_total = (int(billing_ym) // 100) * 12 + (int(billing_ym) % 100)
    candidates = []
    for lag in (1, 2, 0, 3):
        total = billing_total - lag
        y = (total - 1) // 12
        m = (total - 1) % 12 + 1
        if m == month:
            candidates.append((y, m))
    if len(candidates) != 1:
        return None, candidates
    y, m = candidates[0]
    return datetime.date(y, m, day), candidates


def transform(src_path):
    import openpyxl
    wb = openpyxl.load_workbook(src_path, data_only=True)

    entries = []
    skipped = {'blank': 0, 'dummy_header': 0}
    warnings = []

    def add_entry(institution, transaction_date, description, direction, amount,
                  balance, classification, billing_month, card_holder, memo,
                  raw_detail, source_sheet, source_row):
        entries.append({
            'institution': institution,
            'card_holder': card_holder,
            'transaction_date': transaction_date.isoformat(),
            'billing_month': billing_month,
            'description': description,
            'direction': direction,
            'amount': to_number(amount),
            'balance': to_number(balance),
            'classification': classification,
            'memo': memo,
            'raw_detail': raw_detail,
            'source_sheet': source_sheet,
            'source_row': source_row,
        })

    # ---------- 横浜銀行 ----------
    ws = wb['横浜銀行']
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        bunrui, tsuki, dt, shiharai, azukari, kubun, zandaka, tekiyou = row
        if all(v is None for v in row):
            skipped['blank'] += 1
            continue
        if dt is None or tsuki is None:
            warnings.append(('横浜銀行', i, 'date/month missing', row))
            continue
        d = resolve_date_same_year(dt.month, dt.day, tsuki)
        if shiharai is not None:
            direction, amount = '出金', shiharai
        else:
            direction, amount = '入金', azukari
        add_entry('横浜銀行', d, tekiyou, direction, amount, zandaka, bunrui,
                   str(tsuki), None, None, {'取引区分': kubun}, '横浜銀行', i)

    # ---------- 住友銀行 ----------
    ws = wb['住友銀行']
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        bunrui, tsuki, hidzuke_text, hikidashi, azukeire, toriatsukai, zandaka = row
        if all(v is None for v in row):
            skipped['blank'] += 1
            continue
        md = parse_month_day_text(hidzuke_text)
        if md is None or tsuki is None:
            warnings.append(('住友銀行', i, 'date parse failed', row))
            continue
        d = resolve_date_same_year(md[0], md[1], tsuki)
        if hikidashi is not None:
            direction, amount = '出金', hikidashi
        else:
            direction, amount = '入金', azukeire
        add_entry('住友銀行', d, toriatsukai, direction, amount, zandaka, bunrui,
                   str(tsuki), None, None, None, '住友銀行', i)

    # ---------- ゆうちょ ----------
    ws = wb['ゆうちょ']
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        bunrui, tsuki, torihikibi, meisai_id, ukeire, haraidashi, shousai1, shousai2, genzaidaka = row
        if all(v is None for v in row):
            skipped['blank'] += 1
            continue
        if torihikibi is None:
            warnings.append(('ゆうちょ', i, 'date missing', row))
            continue
        s = str(int(torihikibi))
        d = datetime.date(int(s[0:4]), int(s[4:6]), int(s[6:8]))
        if ukeire is not None:
            direction, amount = '入金', ukeire
        else:
            direction, amount = '出金', haraidashi
        desc = shousai1 if not shousai2 else f'{shousai1} {shousai2}'
        add_entry('ゆうちょ', d, desc, direction, amount, genzaidaka, bunrui,
                   str(tsuki) if tsuki else None, None, None,
                   {'入出金明細ID': meisai_id, '詳細1': shousai1, '詳細2': shousai2},
                   'ゆうちょ', i)

    # ---------- みずほ銀行 ----------
    # このシートは2種類の列レイアウトが混在する：
    #   (a) 「残高」列が数値・「お取引内容」列が摘要（119行中80行、想定どおりの並び）
    #   (b) 「残高」列に摘要テキストが入り「お取引内容」列は空（119行中39行、残高データが無い）
    # 前身の「仕訳ルール_統合版.xlsx」設計方針シートが「みずほ銀行80行」とだけ記載していたのは
    # (a)のみを数えていたためと判明（2026-08-22調査）。本スクリプトは(a)(b)両方を取り込む。
    ws = wb['みずほ銀行']
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        bunrui, tsuki, meisai_tsuuban, hidzuke_text, hikidashi, azukeire, zandaka, toriatsukai = row
        if all(v is None for v in row):
            skipped['blank'] += 1
            continue
        md = parse_month_day_text(hidzuke_text)
        if md is None or tsuki is None:
            warnings.append(('みずほ銀行', i, 'date parse failed', row))
            continue
        d = resolve_date_same_year(md[0], md[1], tsuki)
        if hikidashi is not None:
            direction, amount = '出金', hikidashi
        else:
            direction, amount = '入金', azukeire
        if isinstance(zandaka, (int, float)):
            balance, description = zandaka, toriatsukai
        else:
            balance, description = None, zandaka
        add_entry('みずほ銀行', d, description, direction, amount, balance, bunrui,
                   str(tsuki), None, None, {'明細通番': meisai_tsuuban}, 'みずほ銀行', i)

    # ---------- 横浜VISA ----------
    ws = wb['横浜VISA']
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        bunrui, goukei, riyoubi_text, tenmei, kingaku, suuryou, tanka, goukei_kingaku, bikou = row
        if all(v is None for v in row):
            skipped['blank'] += 1
            continue
        full_date = parse_full_date_text(riyoubi_text)
        if full_date is not None:
            d = full_date
        else:
            md = parse_month_day_text(riyoubi_text)
            if md is None or goukei is None:
                warnings.append(('横浜VISA', i, 'date parse failed', row))
                continue
            d, candidates = resolve_date_billing_lag(md[0], md[1], goukei)
            if d is None:
                warnings.append(('横浜VISA', i, 'billing lag ambiguous', row, candidates))
                continue
        # 「金額」列が空でも「合計金額」列に値がある行がある（ポイント交換等のキャッシュバック行、マイナス値）
        effective_amount = kingaku if kingaku is not None else goukei_kingaku
        direction = '出金'
        if effective_amount is not None and to_number(effective_amount) < 0:
            direction = '入金'
            effective_amount = abs(to_number(effective_amount))
        add_entry('横浜VISA', d, tenmei, direction, effective_amount, None, bunrui,
                   str(goukei), None, bikou,
                   {'数量': suuryou, '単価': tanka, '合計金額': goukei_kingaku},
                   '横浜VISA', i)

    # ---------- 住友VISA ----------
    ws = wb['住友VISA']
    card_holder = '智広'  # 支払月が変わるたびに智広様側にリセットされる（デフォルト）
    current_month = None
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        bunrui, bikou, goukei, hidzuke_text, naiyou, kingaku = row[:6]
        if all(v is None for v in row):
            skipped['blank'] += 1
            continue
        if goukei != current_month:
            current_month = goukei
            card_holder = '智広'
        if is_dash_only(bunrui):
            # カード名義ヘッダー行：実取引ではない。名義を切り替えて取り込み対象外にする
            name_field = hidzuke_text or ''
            if '恵美' in str(name_field):
                card_holder = '恵美'
            elif '智広' in str(name_field):
                card_holder = '智広'
            skipped['dummy_header'] += 1
            continue
        md = parse_month_day_text(hidzuke_text)
        if md is None or goukei is None:
            warnings.append(('住友VISA', i, 'date parse failed', row))
            continue
        d, candidates = resolve_date_billing_lag(md[0], md[1], goukei)
        if d is None:
            warnings.append(('住友VISA', i, 'billing lag ambiguous', row, candidates))
            continue
        direction = '出金'
        effective_amount = kingaku
        if effective_amount is not None and to_number(effective_amount) < 0:
            direction = '入金'  # 返品・返金等のマイナス計上
            effective_amount = abs(to_number(effective_amount))
        add_entry('住友VISA', d, naiyou, direction, effective_amount, None, bunrui,
                   str(goukei), card_holder, bikou, None, '住友VISA', i)

    # ---------- 楽天カード ----------
    ws = wb['楽天カード']
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if all(v is None for v in row):
            skipped['blank'] += 1
            continue
        bunrui = row[0]
        tsuki = row[2]
        riyoubi = row[3]
        tenmei = row[4]
        riyousha = row[5]
        shiharai_houhou = row[6]
        kingaku = row[7]
        tesuuryou = row[8]
        shiharai_soukagu = row[9]
        touge_shiharai = row[10]
        kurikoshi = row[11]
        shinki_sign = row[12]
        if tenmei is not None and '現地利用額' in str(tenmei):
            # 海外利用時の換算レート注記行（実取引ではない）
            skipped['dummy_header'] += 1
            continue
        if riyoubi is None or tsuki is None:
            warnings.append(('楽天カード', i, 'date missing', row))
            continue
        d, candidates = resolve_date_billing_lag(riyoubi.month, riyoubi.day, tsuki)
        if d is None:
            warnings.append(('楽天カード', i, 'billing lag ambiguous', row, candidates))
            continue
        direction = '出金'
        effective_amount = kingaku
        if effective_amount is not None and to_number(effective_amount) < 0:
            direction = '入金'
            effective_amount = abs(to_number(effective_amount))
        add_entry('楽天カード', d, tenmei, direction, effective_amount, None, bunrui,
                   str(tsuki), None, None,
                   {'利用者': riyousha, '支払方法': shiharai_houhou, '支払手数料': tesuuryou,
                    '支払総額': shiharai_soukagu, '当月支払金額': touge_shiharai,
                    '繰越残高': kurikoshi, '新規サイン': shinki_sign},
                   '楽天カード', i)

    return entries, skipped, warnings


def main():
    if len(sys.argv) < 2:
        print('使い方: python3 transform.py <入力xlsxのパス> [出力jsonのパス]', file=sys.stderr)
        sys.exit(1)
    src_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else 'journal_entries.json'

    entries, skipped, warnings = transform(src_path)

    print(f'entries: {len(entries)}')
    print(f'skipped: {skipped}')
    print(f'warnings: {len(warnings)}')
    for w in warnings[:30]:
        print(' WARN', w)

    by_inst = {}
    for e in entries:
        by_inst.setdefault(e['institution'], 0)
        by_inst[e['institution']] += 1
    print('by institution:', by_inst)

    bad_amount = [e for e in entries if not isinstance(e['amount'], (int, float)) or e['amount'] < 0]
    bad_desc = [e for e in entries if not e['description']]
    if bad_amount or bad_desc:
        print(f'!! 要確認: bad_amount={len(bad_amount)} bad_desc={len(bad_desc)}', file=sys.stderr)

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(entries, f, ensure_ascii=False)
    print(f'wrote {out_path}')


if __name__ == '__main__':
    main()
