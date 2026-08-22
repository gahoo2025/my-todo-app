#!/usr/bin/env python3
"""
journal_classification_map 移行スクリプト（1/2: 変換）

「仕訳１を行ったリスト.xlsx」の `仕訳ルール２`シート（142行、分類1→分類2→分類3の展開ルール）を
journal_classification_map テーブル用のJSONに変換する。使い方は transform.py と対称。

使い方:
  python3 transform_classification_map.py <入力xlsxのパス> [出力jsonのパス]
"""
import sys
import json


def transform(src_path):
    import openpyxl
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb['仕訳ルール２']

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        bunrui3, torihikisaki, bunrui2, bunrui1, nyushukkin, kanren = row[:6]
        if all(v is None for v in row):
            continue
        rows.append({
            'institution_or_group': torihikisaki,
            'classification_1': bunrui1,
            'classification_2': bunrui2,
            'classification_3': bunrui3,
            'cashflow_direction': nyushukkin,
            'related_group': kanren,
            'status': '既存',
            'note': None,
        })
    return rows


def main():
    if len(sys.argv) < 2:
        print('使い方: python3 transform_classification_map.py <入力xlsxのパス> [出力jsonのパス]', file=sys.stderr)
        sys.exit(1)
    src_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else 'journal_classification_map.json'

    rows = transform(src_path)
    print(f'rows: {len(rows)}')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False)
    print(f'wrote {out_path}')


if __name__ == '__main__':
    main()
